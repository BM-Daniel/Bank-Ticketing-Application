from Customer import Customer
from Teller import TellerQueue
from operations import *
import openpyxl, xlrd, os
from openpyxl import Workbook

<<<<<<< HEAD

def get_user():
    """
    Returns the type of user
    :return: user
    """

    # Validates the user's input
    while True:
        print("User Categories:")
        print("\t1. Admin")
        print("\t2. Customer")
        print("\t3. Teller")
        user = input("Enter the type of user: ")

        if user.title().strip() == "Admin" or user == "1":
            return 1  # Returns 1 for Admin
        elif user.title().strip() == "Customer" or user == "2":
            return 2  # Returns 2 for Customer
        elif user.title().strip() == "Teller" or user == "3":
            return 3  # Returns 3 for Teller
        else:
            print("\nEnter a valid option!!!\n")  # Displays error message


def get_service():
    """
    Returns the type of service
    :return: service
    """

    # Validates the user's input
    while True:
        print("Services:")
        print("\t1. Deposit")
        print("\t2. Withdrawal")
        print("\t3. Transfer")
        service_request = input("Select a service: ")

        if service_request.title().strip() == "Deposit" or service_request == "1":
            return "Deposit"
        elif service_request.title().strip() == "Withdrawal"  or service_request == "2":
            return "Withdrawal"
        elif service_request.title().strip() == "Transfer" or service_request == "3" :
            return "Transfer"
        else:
            print("\nEnter a valid option!!!\n")  # Displays error message


def get_admin_task():
    """
    Returns the task of admin
    :return: task
    """

    # Validates admin input
    while True:
        try:
            print("\nWelcome Admin!")
            input("Enter password: ")
            print("\nTasks: ")
            print("\t1. Create a teller queue")
            print("\t2. Reassign active teller")
            task = int(input("Enter task: "))

            if task < 1 or task > 2:
                print("\nEnter a correct task!\n")
            else:
                return task
        except ValueError:
            print("\nEnter a correct task!")


def admin(teller_id_no, teller_list):
=======
def admin(teller_list):
>>>>>>> c46e2567d44b4b606f4cd83c3b6c8c77f0d54e7c
    """
    Allows admin to create teller queues and re-assign an active teller
    :param teller_id:
    :param teller_list:
    :return:
    """
    task = get_admin_task()

    # Creates a new teller queue
    if task == 1:
        print("\nPlease input teller details:")
        teller_name = input("Name: ").title()
        teller_service = get_service()

<<<<<<< HEAD
        teller_id = teller_name[0 : 2].lower() + str(teller_id_no)  # Generates a new teller id
=======
        teller_id = teller_name[0:4].lower() + "21"  # Generates a new teller id
>>>>>>> c46e2567d44b4b606f4cd83c3b6c8c77f0d54e7c

        if teller_service == "Deposit":
            deposit = TellerQueue(teller_id, teller_name, teller_service)
            teller_list.append(deposit)  # Adds deposit queue to the teller list
        elif teller_service == "Withdrawal":
            withdrawal = TellerQueue(teller_id, teller_name, teller_service)
            teller_list.append(withdrawal)  # Adds withdrawal queue to the teller list
        else:
            transfer = TellerQueue(teller_id, teller_name, teller_service)
            teller_list.append(transfer)  # Adds transfer queue to the teller list

        print(f"\n{teller_service} queue has been created\n")
        return
<<<<<<< HEAD

    else:
        # Allows the admin to change the active teller
        print("\nPlease enter the service queue you want to edit") 
        service = get_service()  

        if len(teller_list) == 0:
            print("\nNo teller queues have been created\n")
            return

        for i in range(len(teller_list)):
            if teller_list[i].service == service:
                new_teller_name = input("Enter the new teller's name: ")
                teller_list[i].teller_name = new_teller_name.title()
                return f"{new_teller_name} has been reassigned to {service} queue"
            else:
                print("\nThe service queue does not exist!\n")
=======
>>>>>>> c46e2567d44b4b606f4cd83c3b6c8c77f0d54e7c

    else:
        # Allows the admin to change the active teller
        print("\nPlease enter the service queue you want to edit") 
        service = get_service()  

<<<<<<< HEAD

def customer(ticket_id, teller_list):
=======
        if len(teller_list) == 0:
            print("\nNo teller queues have been created\n")
            return

        teller_absent = True

        for i in range(len(teller_list)):
            if teller_list[i].service == service:
                new_teller_name = input("Enter the new teller's name: ")
                teller_list[i].teller_name = new_teller_name.title()
                print(f"{new_teller_name} has been reassigned to {service} queue\n")
                teller_absent  = False

        if teller_absent:
            print("\nThe service queue does not exist!\n")

def customer(teller_list, sheet):
>>>>>>> c46e2567d44b4b606f4cd83c3b6c8c77f0d54e7c
    """
    Adds a new customer to a particular teller queue
    :param ticket_id:
    :param teller_queue:
    :return: success message
    """
    if len(teller_list) == 0:
        print("\nNo service queue available. Contact the admin\n")
        return

    print("\nPlease enter the following details: ")
    customer_name = input("Name: ").title()
    service_request = get_service()

    service_present = False

    for i in range(len(teller_list)):
        if service_request == teller_list[i].service:
            service_present = True

    if not service_present:
        print(f"\n{service_request} queue is unavailable. Please contact admin\n")
        return

    amount = 0
    while True:
        try:
            amount = round(float(input("Amount of Transaction: GH¢")), 2)
            if amount > 0:
                break
            else:
                print("Enter a valid amount")
        except ValueError:
            print("Enter a valid amount")

    priority_level = False
    if amount > 10000:
        priority_level = True

    if priority_level:
        ticket_id = make_account_id(prefix='PC-')# Generates a new ticket id
    else:
        ticket_id = make_account_id(prefix='NC-')

    new_customer = Customer(customer_name, service_request, ticket_id, priority_level)

<<<<<<< HEAD
    service_present = False

    for i in range(len(teller_list)):
        if service_request == teller_list[i].service:
            service_present = True

    if service_present:
        for i in range(len(teller_list)):
            teller_list[i].enqueue(new_customer)  # Adds a customer to a particular teller queue
    else:
        print(f"\n{service_request} queue is unavailable. Please contact admin\n")
        return

    print()
    print("*"*10, "Welcome", "*"*10)
=======
    customer_tuple = (customer_name, service_request, ticket_id, priority_level, "¢" + str(amount))

    sheet.append(customer_tuple)

    for i in range(len(teller_list)):
        teller_list[i].enqueue(new_customer)  # Adds a customer to a particular teller queue

    print()
    print("*"*15, "Welcome", "*"*15)
>>>>>>> c46e2567d44b4b606f4cd83c3b6c8c77f0d54e7c
    print("Ticket Details: ")
    print(f"\t\tName: {customer_name}")
    print(f"\t\tTicket Number: #{ticket_id}")
    print(f"\t\tTransaction: {service_request}")
<<<<<<< HEAD
    print("*"*29, "\n")

def get_choice():
    while True:
        try:
            print("\t1. Serve Queue")
            print("\t2. View Queue")
            print("\t3. Log out")
            choice = int(input("Enter choice: "))

            if choice < 1 or choice > 3:
                print("\nEnter a correct choice\n")
            else:
                return choice
        except ValueError:
            print("\nEnter a valid number\n")



def validate_teller(teller_list):
    correct = True
    while True:
        print("\nPlease enter the following details: ")
        teller_id = input("Teller id: ")
        print("Enter service type: ")
        service_request = get_service()
        print()

        for i in range(len(teller_list)):
            if teller_id == teller_list[i].teller_ID and service_request == teller_list[i].service:
                return i
            else:
                correct = False

        if not correct:
            print("Incorrect Details!!!")

=======
    print("*"*38, "\n")
>>>>>>> c46e2567d44b4b606f4cd83c3b6c8c77f0d54e7c

def teller(teller_list):
    if len(teller_list) == 0:
        print("\nNo teller queue. Contact the admin\n")
        return

    teller_index = validate_teller(teller_list)

    if teller_list[teller_index].is_empty():
        print("No customer available\n")
        return

    while True:
<<<<<<< HEAD
        choice = get_choice()
=======
        choice = get_teller_choice()
>>>>>>> c46e2567d44b4b606f4cd83c3b6c8c77f0d54e7c

        if choice == 3:
            print("\n..........\n")
            break
        elif choice == 2:
            print()
            print(f"{teller_list[teller_index]}")
        else:
            print()
            print(teller_list[teller_index].dequeue())
            print(f"{teller_list[teller_index]}")

            if teller_list[teller_index].is_empty():
                print("No more customers in this queue.\n")
                break

def main():
    teller_list = []
<<<<<<< HEAD
    ticket_id = 0
    teller_id = 21
=======

    dest_filename = 'Customers.xlsx'

    if os.path.isfile(dest_filename):
        try:
            file = openpyxl.load_workbook(dest_filename)
        except PermissionError:
            print("\nFile is open. Close it")
    else:
        file = Workbook()
        sheet = file.active

        heading = ("Name", "Service Request", "Ticket ID", "Priority", "Amount")

        for i in range(1, 6):
            sheet.cell(row=1, column=i).value = heading[i - 1]

        file.save(dest_filename)

    sheet = file.active
>>>>>>> c46e2567d44b4b606f4cd83c3b6c8c77f0d54e7c

    print("*" * 6, "Welcome to DIGESJC Bank", "*" * 6, end="\n")  # Welcomes user

    while True:
        user = get_user()

        if user == 1:
<<<<<<< HEAD
            admin(teller_id, teller_list)
        elif user == 2:
            customer(ticket_id, teller_list)
            ticket_id += 1
        else:
            teller(teller_list)

=======
            admin(teller_list)
        elif user == 2:
            customer(teller_list, sheet)
        else:
            teller(teller_list)

        file.save(dest_filename)



>>>>>>> c46e2567d44b4b606f4cd83c3b6c8c77f0d54e7c
main()